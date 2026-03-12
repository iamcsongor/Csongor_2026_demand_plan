"""
Cambri Demand Plan — GitHub Actions compute script
----------------------------------------------------
Downloads the SharePoint Excel file, reads the manually-maintained
'Actuals' summary sheet AND the raw 'Demand Plan Actuals' sheet,
then writes data.json to the repo root.

Run:
    pip install requests openpyxl
    python compute.py
"""

import datetime
import json
from io import BytesIO

import requests
from openpyxl import load_workbook

# ── CONFIG ────────────────────────────────────────────────────────────────────
SHAREPOINT_URL = (
    "https://wiseandsallycom-my.sharepoint.com/:x:/g/personal/"
    "csongor_doma_cambri_io/"
    "IQDKVYMS8mOJS5WxpQi7eylEAWeqX1IwmUSiOAj2kDGFOns"
    "?e=Zh70hH&download=1"
)
TARGET_YEAR     = 2026
OUTPUT_FILE     = "data.json"
SF_BASE_URL     = "https://cambri.lightning.force.com/lightning/r/Opportunity/{id}/view"

# ── FIXED ANNUAL TARGETS ──────────────────────────────────────────────────────
TARGETS = {
    "marketing": {"s1": 130,  "s2": 104, "poc": 36.4, "wins": 18.2, "value": 2275000},
    "bdr":       {"s1": 313,  "s2": 72,  "poc": 19.4, "wins": 9.7,  "value": 730000 },
    "partner":   {"s1": 50,   "s2": 22,  "poc": 11,   "wins": 5.5,  "value": 550000 },
}

SRC_MAP = {
    "marketing":  "marketing",
    "bdr":        "bdr",
    "partner":    "partner",
}

METRIC_MAP = {
    "s1":                   "s1",
    "opps created":         "s1",
    "opportunities created":"s1",
    "s2":                   "s2",
    "advanced to s2":       "s2",
    "poc":                  "poc",
    "wins #":               "wins",
    "wins#":                "wins",
    "wins number":          "wins",
    "wins €":               "value",
    "wins€":                "value",
    "value":                "value",
    "wins value":           "value",
    "bookings":             "value",
}

# Contribution Channel → normalized channel key
CHANNEL_MAP = {
    "marketing sourced": "marketing",
    "bdr sourced":       "bdr",
    "sales sourced":     "bdr",
    "partner sourced":   "partner",
    "csm sourced":       "partner",
}


def download_workbook():
    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0 Safari/537.36"
    )
    resp = session.get(SHAREPOINT_URL, allow_redirects=True, timeout=60)
    resp.raise_for_status()
    return load_workbook(BytesIO(resp.content), data_only=True)


def read_arr_sheet(wb):
    """
    Read the 'ARR' tab.  Finds columns by header name (row 2) so the sheet
    column order doesn't matter.
    Returns: {
        "monthly_target":  [float|None]*12,
        "monthly_actual":  [float|None]*12,
        "reforecast":      [float|None]*12,   # cumulative Re-forecast column
    }
    """
    empty = {
        "monthly_target": [None]*12,
        "monthly_actual": [None]*12,
        "reforecast":     [None]*12,
    }
    if "ARR" not in wb.sheetnames:
        return empty

    ws = wb["ARR"]

    def _f(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    # --- find header row (row 2) and map column names → 0-based index ---
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    col_idx = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip().lower()
        col_idx[key] = i

    def _col(name):
        """Return 0-based index for a header, searching by substring."""
        for k, v in col_idx.items():
            if name in k:
                return v
        return None

    i_target     = _col("target")
    i_actual     = _col("actual")
    i_reforecast = _col("re-forecast") or _col("reforecast") or _col("re forecast")

    monthly_target  = []
    monthly_actual  = []
    reforecast      = []

    for row in ws.iter_rows(min_row=3, max_row=14, values_only=True):
        monthly_target.append( _f(row[i_target])     if i_target     is not None and i_target     < len(row) else None)
        monthly_actual.append( _f(row[i_actual])     if i_actual     is not None and i_actual     < len(row) else None)
        reforecast.append(     _f(row[i_reforecast]) if i_reforecast is not None and i_reforecast < len(row) else None)

    # Pad to 12
    for lst in (monthly_target, monthly_actual, reforecast):
        while len(lst) < 12:
            lst.append(None)

    return {
        "monthly_target": monthly_target,
        "monthly_actual": monthly_actual,
        "reforecast":     reforecast,
    }


def read_actuals_sheet(wb):
    """
    Read the 'Actuals' summary sheet.
    Row 3: headers (Source / Metric / Jan–Dec)
    Row 4+: data rows
    """
    ws = wb["Actuals"]
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(all_rows):
        vals = [str(v).strip().lower() if v is not None else "" for v in row[:3]]
        if "source" in vals or "metric" in vals:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row in 'Actuals' sheet")

    data_rows = all_rows[header_idx + 1:]

    actuals = {
        src: {m: [None] * 12 for m in ("s1", "s2", "poc", "wins", "value")}
        for src in ("marketing", "bdr", "partner")
    }

    current_src = None

    for row in data_rows:
        if all(v is None for v in row):
            continue

        src_raw    = str(row[0]).strip().lower() if row[0] is not None else ""
        metric_raw = str(row[1]).strip().lower() if row[1] is not None else ""

        if src_raw in SRC_MAP:
            current_src = SRC_MAP[src_raw]

        if current_src is None:
            continue

        metric = METRIC_MAP.get(metric_raw)
        if metric is None:
            continue

        for mo_idx in range(12):
            col_idx = 2 + mo_idx
            if col_idx >= len(row):
                break
            val = row[col_idx]
            if val is not None:
                try:
                    actuals[current_src][metric][mo_idx] = float(val)
                except (TypeError, ValueError):
                    pass

    return actuals


def _month_from_excel_serial(v):
    """
    Excel MONTH() formulas stored in date-formatted cells come through as
    datetime(1900, 1, N) where N is the month number 1–12.
    Plain integers come through as-is.
    """
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        # 1900-era = the day IS the month number
        return v.day if v.year == 1900 else v.month
    if isinstance(v, (int, float)):
        n = int(v)
        return n if 1 <= n <= 12 else None
    return None


def _fmt_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)


def read_last_sf_sync(wb):
    """
    Return the most recent datetime from 'Automatic Operations Events Log'
    column A (skipping the header row) as an ISO-8601 string, or None.
    """
    if "Automatic Operations Events Log" not in wb.sheetnames:
        return None
    ws = wb["Automatic Operations Events Log"]
    latest = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0] if row else None
        if isinstance(val, datetime.datetime):
            if latest is None or val > latest:
                latest = val
    return latest.isoformat() if latest else None


def read_deals(wb):
    """
    Read individual opportunity records from 'Demand Plan Actuals'.
    Returns a list of deal dicts, each tagged with which metrics/months
    (0-indexed) they contribute to in TARGET_YEAR.
    """
    ws = wb["Demand Plan Actuals"]

    # Find the header row (it contains 'Opportunity Owner')
    headers = None
    header_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "Opportunity Owner":
            headers = list(row)
            header_row_idx = row_idx
            break

    if headers is None:
        print("[compute.py] Warning: could not find deal header row — skipping deals")
        return []

    H = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, col_name):
        idx = H.get(col_name)
        return row[idx] if idx is not None and idx < len(row) else None

    deals = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue

        opp_id   = get(row, "Opp Casesafe ID 18")
        opp_name = get(row, "Opportunity Name")
        if not opp_id or not opp_name:
            continue

        owner      = get(row, "Opportunity Owner")
        created_by = get(row, "Created By")
        amount     = get(row, "Amount (converted)")
        close_date = get(row, "Close Date")
        opp_type   = get(row, "Type")
        rec_type   = get(row, "Opportunity Record Type")
        stage      = get(row, "Stage")
        channel_raw = get(row, "Contribution Channel")
        channel    = CHANNEL_MAP.get(str(channel_raw).strip().lower() if channel_raw else "", None)

        created_year = get(row, "Created YEAR")
        created_m_raw = get(row, "Created Month2")
        s2_year      = get(row, "S2 Year")
        s2_m_raw     = get(row, "S2 Month")
        close_year   = get(row, "Close Date YEAR")
        close_month  = get(row, "Close Date Month")
        fsr_year     = get(row, "FSR YEAR")
        fsr_m_raw    = get(row, "FSR MONTH")

        created_month = _month_from_excel_serial(created_m_raw)
        s2_month      = _month_from_excel_serial(s2_m_raw)
        fsr_month     = _month_from_excel_serial(fsr_m_raw)

        # Which metrics/months does this deal contribute to in TARGET_YEAR?
        s1_months   = []
        s2_months   = []
        poc_months  = []
        wins_months = []

        if created_year == TARGET_YEAR and created_month and 1 <= created_month <= 12:
            s1_months.append(created_month - 1)  # 0-indexed

        if s2_year == TARGET_YEAR and s2_month and 1 <= s2_month <= 12:
            s2_months.append(s2_month - 1)

        # POC: tracked via FSR month when available, fallback not yet defined
        if fsr_year == TARGET_YEAR and fsr_month and 1 <= fsr_month <= 12:
            poc_months.append(fsr_month - 1)

        if (close_year == TARGET_YEAR and stage == "Closed Won"
                and close_month and 1 <= int(close_month) <= 12):
            wins_months.append(int(close_month) - 1)

        # Only include deals relevant to TARGET_YEAR
        if not any([s1_months, s2_months, poc_months, wins_months]):
            continue

        deals.append({
            "id":          opp_id,
            "sf_url":      SF_BASE_URL.format(id=opp_id),
            "owner":       owner,
            "created_by":  created_by,
            "name":        opp_name,
            "amount":      float(amount) if amount is not None else None,
            "close_date":  _fmt_date(close_date),
            "type":        opp_type,
            "record_type": rec_type,
            "channel":     channel,
            "channel_raw": channel_raw,
            "stage":       stage,
            "s1_months":   s1_months,
            "s2_months":   s2_months,
            "poc_months":  poc_months,
            "wins_months": wins_months,
        })

    return deals


def main():
    print("[compute.py] Downloading SharePoint file…")
    wb = download_workbook()

    print("[compute.py] Reading Actuals sheet…")
    actuals = read_actuals_sheet(wb)

    print("[compute.py] Reading ARR sheet…")
    arr = read_arr_sheet(wb)

    print("[compute.py] Reading deal-level data…")
    deals = read_deals(wb)

    print("[compute.py] Reading last SF sync date…")
    last_sf_sync = read_last_sf_sync(wb)
    wb.close()

    payload = {
        "targets":       TARGETS,
        "actuals":       actuals,
        "arr":           arr,
        "deals":         deals,
        "last_sf_sync":  last_sf_sync,
        "last_fetched":  datetime.datetime.utcnow().isoformat() + "Z",
        "source":        "sharepoint_actuals_sheet",
    }

    with open(OUTPUT_FILE, "w") as f:
        json.dump(payload, f, indent=2)

    print(f"[compute.py] ✓ Wrote {OUTPUT_FILE} ({len(deals)} deals)")

    for src in ("marketing", "bdr", "partner"):
        s1 = actuals[src]["s1"][:3]
        print(f"  {src:12s} S1 Jan-Mar: {s1}")


if __name__ == "__main__":
    main()
