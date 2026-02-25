"""
Cambri Demand Plan — Live Data Server
--------------------------------------
Fetches the SharePoint Excel file, computes the 5 demand-plan metrics
from the raw Salesforce export sheet, and serves them as JSON on
http://localhost:5050/data

The dashboard polls this endpoint; no Google Sheets or auth needed.

Run:
    pip install flask requests openpyxl
    python server.py
"""

import datetime
import threading
import time
from collections import defaultdict
from io import BytesIO

import requests
from flask import Flask, jsonify
from openpyxl import load_workbook

# ── CONFIG ────────────────────────────────────────────────────────────────────
SHAREPOINT_URL = (
    "https://wiseandsallycom-my.sharepoint.com/:x:/g/personal/"
    "csongor_doma_cambri_io/"
    "IQDKVYMS8mOJS5WxpQi7eylEAWeqX1IwmUSiOAj2kDGFOns"
    "?e=Zh70hH&download=1"
)
REFRESH_INTERVAL = 300   # seconds between background refreshes (5 min)
PORT             = 5050
TARGET_YEAR      = 2026

# ── FIXED ANNUAL TARGETS ──────────────────────────────────────────────────────
TARGETS = {
    "marketing": {"s1": 130,  "s2": 104, "poc": 36.4, "wins": 18.2, "value": 2275000},
    "bdr":       {"s1": 313,  "s2": 72,  "poc": 19.4, "wins": 9.7,  "value": 730000 },
    "partner":   {"s1": 50,   "s2": 22,  "poc": 11,   "wins": 5.5,  "value": 550000 },
}

CH_MAP = {
    "Marketing Sourced": "marketing",
    "BDR Sourced":       "bdr",
    "Partner Sourced":   "partner",
}

# ── STATE ─────────────────────────────────────────────────────────────────────
_cache       = {}          # latest computed payload
_last_fetched = None
_lock        = threading.Lock()


def _download_workbook():
    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0 Safari/537.36"
    )
    resp = session.get(SHAREPOINT_URL, allow_redirects=True, timeout=30)
    resp.raise_for_status()
    return load_workbook(BytesIO(resp.content), data_only=True, read_only=True)


def _safe_int_month(val):
    """Return 1-12 int from an int or datetime month field, else None."""
    if isinstance(val, int) and 1 <= val <= 12:
        return val
    if isinstance(val, datetime.datetime):
        return val.day if 1 <= val.day <= 12 else None
    return None


def _compute_actuals(wb):
    ws = wb["Demand Plan Actuals"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Header row is row 5 (index 4) — first two rows blank, row 3 = timestamp
    header_row = None
    for i, row in enumerate(all_rows):
        if row[0] == "Opportunity Owner":
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find header row in 'Demand Plan Actuals'")

    headers = list(all_rows[header_row])
    col     = {h: i for i, h in enumerate(headers) if h is not None}
    data    = [
        all_rows[i]
        for i in range(header_row + 1, len(all_rows))
        if any(v is not None for v in all_rows[i])
    ]

    # Zero-filled result structure: [jan … dec]
    actuals = {
        src: {m: [None] * 12 for m in ("s1", "s2", "poc", "wins", "value")}
        for src in ("marketing", "bdr", "partner")
    }

    def inc(src, metric, mo_idx, amount=1):
        cur = actuals[src][metric][mo_idx]
        actuals[src][metric][mo_idx] = (cur or 0) + amount

    for row in data:
        src = CH_MAP.get(str(row[col["Contribution Channel"]]))
        if not src:
            continue

        cr_year = row[col["Created YEAR"]]
        cr_mo   = _safe_int_month(row[col["Created Month2"]])
        s2_year = row[col["S2 Year"]]
        s2_mo   = _safe_int_month(row[col["S2 Month"]])
        close_d = row[col["Close Date"]]
        stage   = row[col["Stage"]]
        poc_st  = row[col["POC Status c"]]
        amount  = float(row[col["Amount (converted)"]] or 0)

        # S1: created this year
        if cr_year == TARGET_YEAR and cr_mo:
            inc(src, "s1", cr_mo - 1)

        # S2: advanced to S2 this year
        if s2_year == TARGET_YEAR and s2_mo:
            inc(src, "s2", s2_mo - 1)

        # POC: any record where POC Status is set, created this year
        if poc_st is not None and cr_year == TARGET_YEAR and cr_mo:
            inc(src, "poc", cr_mo - 1)

        # Wins # and Wins €: Closed Won with close date in this year
        if (
            stage == "Closed Won"
            and isinstance(close_d, datetime.datetime)
            and close_d.year == TARGET_YEAR
        ):
            mo_idx = close_d.month - 1
            inc(src, "wins",  mo_idx)
            inc(src, "value", mo_idx, amount)

    return actuals


def fetch_and_compute():
    global _cache, _last_fetched
    try:
        wb      = _download_workbook()
        actuals = _compute_actuals(wb)
        wb.close()
        payload = {
            "targets":       TARGETS,
            "actuals":       actuals,
            "last_fetched":  datetime.datetime.utcnow().isoformat() + "Z",
            "source":        "sharepoint_live",
        }
        with _lock:
            _cache       = payload
            _last_fetched = datetime.datetime.utcnow()
        print(f"[{datetime.datetime.now():%H:%M:%S}] Data refreshed from SharePoint ✓")
        return True
    except Exception as e:
        print(f"[{datetime.datetime.now():%H:%M:%S}] Refresh failed: {e}")
        return False


def _background_refresh():
    while True:
        time.sleep(REFRESH_INTERVAL)
        fetch_and_compute()


# ── FLASK APP ─────────────────────────────────────────────────────────────────
app = Flask(__name__)


@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return response


@app.route("/data")
def data():
    with _lock:
        if not _cache:
            return jsonify({"error": "Data not yet loaded"}), 503
        return jsonify(_cache)


@app.route("/refresh", methods=["POST", "GET"])
def manual_refresh():
    ok = fetch_and_compute()
    return jsonify({"ok": ok, "last_fetched": _cache.get("last_fetched")})


@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "last_fetched": _last_fetched.isoformat() if _last_fetched else None,
    })


# ── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print(" Cambri Demand Plan Server")
    print(f" Fetching from SharePoint …")
    ok = fetch_and_compute()
    if not ok:
        print(" ⚠  Initial fetch failed — will retry in background")

    t = threading.Thread(target=_background_refresh, daemon=True)
    t.start()

    print(f" Serving on http://localhost:{PORT}")
    print(f" Dashboard: open demand_dashboard.html in your browser")
    print("=" * 55)
    app.run(host="0.0.0.0", port=PORT, debug=False, threaded=True)
